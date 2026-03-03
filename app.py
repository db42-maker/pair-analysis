import os
import io
import time
from flask import Flask, jsonify, send_from_directory, request, send_file
from flask_cors import CORS
import numpy as np
import requests as req_lib

app = Flask(__name__, static_folder=".")
CORS(app)

ALPHA_VANTAGE_KEY = os.environ.get("ALPHA_VANTAGE_KEY", "")


def fetch_daily_prices(ticker, api_key):
    """Fetch daily adjusted close prices from Alpha Vantage."""
    url = (
        f"https://www.alphavantage.co/query"
        f"?function=TIME_SERIES_DAILY_ADJUSTED"
        f"&symbol={ticker}"
        f"&outputsize=full"
        f"&apikey={api_key}"
    )
    r = req_lib.get(url, timeout=15)
    data = r.json()

    if "Error Message" in data:
        raise ValueError(f"Ticker not found: {ticker}")
    if "Note" in data:
        raise ValueError("Alpha Vantage rate limit hit. Please wait a minute and try again.")
    if "Information" in data:
        raise ValueError("Alpha Vantage API limit reached. Please try again later.")

    ts = data.get("Time Series (Daily)", {})
    if not ts:
        raise ValueError(f"No data returned for {ticker}")

    # Sort dates ascending
    sorted_dates = sorted(ts.keys())
    dates = sorted_dates
    prices = [float(ts[d]["5. adjusted close"]) for d in sorted_dates]
    return dates, prices


def filter_by_period(dates, prices, period):
    """Filter dates/prices to the requested period."""
    from datetime import datetime, timedelta
    period_days = {
        "1mo": 30, "3mo": 90, "6mo": 180,
        "1y": 365, "2y": 730, "5y": 1825
    }
    days = period_days.get(period, 365)
    cutoff = datetime.today() - timedelta(days=days)
    filtered = [(d, p) for d, p in zip(dates, prices) if datetime.strptime(d, "%Y-%m-%d") >= cutoff]
    if not filtered:
        return [], []
    fdates, fprices = zip(*filtered)
    return list(fdates), list(fprices)


def align_series(dates_a, prices_a, dates_b, prices_b):
    """Keep only dates present in both series."""
    map_a = dict(zip(dates_a, prices_a))
    map_b = dict(zip(dates_b, prices_b))
    common = sorted(set(map_a.keys()) & set(map_b.keys()))
    return common, [map_a[d] for d in common], [map_b[d] for d in common]


def align_three(dates_a, prices_a, dates_b, prices_b, dates_c, prices_c):
    map_a = dict(zip(dates_a, prices_a))
    map_b = dict(zip(dates_b, prices_b))
    map_c = dict(zip(dates_c, prices_c))
    common = sorted(set(map_a.keys()) & set(map_b.keys()) & set(map_c.keys()))
    return common, [map_a[d] for d in common], [map_b[d] for d in common], [map_c[d] for d in common]


def daily_returns(prices):
    arr = np.array(prices, dtype=float)
    return np.diff(arr) / arr[:-1]


def compute_stats(prices_a, prices_b, ticker_a, ticker_b):
    ret_a = daily_returns(prices_a)
    ret_b = daily_returns(prices_b)

    corr = float(np.corrcoef(ret_a, ret_b)[0, 1])
    r2 = corr ** 2

    cov = float(np.cov(ret_a, ret_b)[0, 1])
    var_b = float(np.var(ret_b, ddof=1))
    beta = cov / var_b if var_b != 0 else 0.0

    vol_a = float(np.std(ret_a, ddof=1) * np.sqrt(252))
    vol_b = float(np.std(ret_b, ddof=1) * np.sqrt(252))

    spread_returns = ret_a - ret_b
    tracking_error = float(np.std(spread_returns, ddof=1) * np.sqrt(252))

    total_ret_a = float((prices_a[-1] - prices_a[0]) / prices_a[0])
    total_ret_b = float((prices_b[-1] - prices_b[0]) / prices_b[0])

    def sharpe(rets):
        ann_ret = float(np.mean(rets) * 252)
        ann_vol = float(np.std(rets, ddof=1) * np.sqrt(252))
        return ann_ret / ann_vol if ann_vol != 0 else 0.0

    def max_drawdown(prices):
        arr = np.array(prices, dtype=float)
        peak = np.maximum.accumulate(arr)
        dd = (peak - arr) / peak
        return float(np.max(dd))

    return {
        "correlation": round(corr, 6),
        "r2": round(r2, 6),
        "beta": round(beta, 4),
        "tracking_error": round(tracking_error, 6),
        "ticker_a": {
            "symbol": ticker_a,
            "total_return": round(total_ret_a, 6),
            "ann_vol": round(vol_a, 6),
            "sharpe": round(sharpe(ret_a), 4),
            "max_drawdown": round(max_drawdown(prices_a), 6),
            "best_day": round(float(np.max(ret_a)), 6),
            "worst_day": round(float(np.min(ret_a)), 6),
        },
        "ticker_b": {
            "symbol": ticker_b,
            "total_return": round(total_ret_b, 6),
            "ann_vol": round(vol_b, 6),
            "sharpe": round(sharpe(ret_b), 4),
            "max_drawdown": round(max_drawdown(prices_b), 6),
            "best_day": round(float(np.max(ret_b)), 6),
            "worst_day": round(float(np.min(ret_b)), 6),
        },
    }


def compute_rolling_correlation(prices_a, prices_b, window=30):
    ret_a = daily_returns(prices_a)
    ret_b = daily_returns(prices_b)
    n = len(ret_a)
    rolling = []
    for i in range(n):
        if i < window - 1:
            rolling.append(None)
        else:
            a_slice = ret_a[i - window + 1: i + 1]
            b_slice = ret_b[i - window + 1: i + 1]
            corr = float(np.corrcoef(a_slice, b_slice)[0, 1])
            rolling.append(round(corr, 4))
    return rolling


@app.route("/api/compare")
def compare():
    ticker_a = request.args.get("a", "").upper().strip()
    ticker_b = request.args.get("b", "").upper().strip()
    period = request.args.get("period", "1y")
    window = int(request.args.get("window", 30))
    api_key = ALPHA_VANTAGE_KEY

    allowed_periods = {"1mo", "3mo", "6mo", "1y", "2y", "5y"}
    if not ticker_a or not ticker_b:
        return jsonify({"error": "Both tickers are required."}), 400
    if ticker_a == ticker_b:
        return jsonify({"error": "Please enter two different tickers."}), 400
    if period not in allowed_periods:
        return jsonify({"error": "Invalid period."}), 400
    if not api_key:
        return jsonify({"error": "API key not configured on server."}), 500

    try:
        # Fetch all three tickers — small delay to respect rate limits
        dates_a, prices_a_full = fetch_daily_prices(ticker_a, api_key)
        time.sleep(0.5)
        dates_b, prices_b_full = fetch_daily_prices(ticker_b, api_key)
        time.sleep(0.5)
        dates_spy, prices_spy_full = fetch_daily_prices("SPY", api_key)

        # Filter to period
        dates_a, prices_a_full = filter_by_period(dates_a, prices_a_full, period)
        dates_b, prices_b_full = filter_by_period(dates_b, prices_b_full, period)
        dates_spy, prices_spy_full = filter_by_period(dates_spy, prices_spy_full, period)

        # Align all three
        dates, prices_a, prices_b, prices_spy = align_three(
            dates_a, prices_a_full,
            dates_b, prices_b_full,
            dates_spy, prices_spy_full
        )

        if len(dates) < 10:
            return jsonify({"error": "Not enough overlapping data. Check your tickers."}), 400

        # Format dates for display
        from datetime import datetime
        display_dates = [datetime.strptime(d, "%Y-%m-%d").strftime("%b %d '%y") for d in dates]

        idx_a   = [round(p / prices_a[0] * 100, 4) for p in prices_a]
        idx_b   = [round(p / prices_b[0] * 100, 4) for p in prices_b]
        idx_spy = [round(p / prices_spy[0] * 100, 4) for p in prices_spy]

        spread = [round(a / b, 6) for a, b in zip(prices_a, prices_b)]

        ret_a = [round((prices_a[i] - prices_a[i-1]) / prices_a[i-1] * 100, 4) for i in range(1, len(prices_a))]
        ret_b = [round((prices_b[i] - prices_b[i-1]) / prices_b[i-1] * 100, 4) for i in range(1, len(prices_b))]

        rolling_corr = compute_rolling_correlation(prices_a, prices_b, window=window)
        stats = compute_stats(prices_a, prices_b, ticker_a, ticker_b)

        return jsonify({
            "dates": display_dates,
            "indexed_a": idx_a,
            "indexed_b": idx_b,
            "indexed_spy": idx_spy,
            "spread": spread,
            "returns_a": ret_a,
            "returns_b": ret_b,
            "rolling_corr": rolling_corr,
            "rolling_window": window,
            "stats": stats,
            "prices_a": [round(p, 4) for p in prices_a],
            "prices_b": [round(p, 4) for p in prices_b],
            "prices_spy": [round(p, 4) for p in prices_spy],
        })

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Data fetch failed: {str(e)}"}), 500


@app.route("/api/export")
def export():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl not installed."}), 500

    ticker_a = request.args.get("a", "").upper().strip()
    ticker_b = request.args.get("b", "").upper().strip()
    period = request.args.get("period", "1y")
    window = int(request.args.get("window", 30))
    api_key = ALPHA_VANTAGE_KEY

    allowed_periods = {"1mo", "3mo", "6mo", "1y", "2y", "5y"}
    if not ticker_a or not ticker_b or period not in allowed_periods:
        return jsonify({"error": "Invalid parameters."}), 400

    try:
        dates_a, prices_a_full = fetch_daily_prices(ticker_a, api_key)
        time.sleep(0.5)
        dates_b, prices_b_full = fetch_daily_prices(ticker_b, api_key)
        time.sleep(0.5)
        dates_spy, prices_spy_full = fetch_daily_prices("SPY", api_key)

        dates_a, prices_a_full = filter_by_period(dates_a, prices_a_full, period)
        dates_b, prices_b_full = filter_by_period(dates_b, prices_b_full, period)
        dates_spy, prices_spy_full = filter_by_period(dates_spy, prices_spy_full, period)

        dates, prices_a, prices_b, prices_spy = align_three(
            dates_a, prices_a_full,
            dates_b, prices_b_full,
            dates_spy, prices_spy_full
        )

        rolling_corr = compute_rolling_correlation(prices_a, prices_b, window=window)
        stats = compute_stats(prices_a, prices_b, ticker_a, ticker_b)

        wb = openpyxl.Workbook()

        dark_fill   = PatternFill("solid", fgColor="0A0C10")
        header_fill = PatternFill("solid", fgColor="C8A96E")
        alt_fill    = PatternFill("solid", fgColor="111318")
        white_font  = Font(color="E8E4DC", name="Courier New", size=10)
        dark_font   = Font(color="0A0C10", bold=True, name="Courier New", size=10)
        gold_font   = Font(color="C8A96E", bold=True, name="Courier New", size=12)
        muted_font  = Font(color="6B7280", name="Courier New", size=9)
        center      = Alignment(horizontal="center", vertical="center")
        left        = Alignment(horizontal="left", vertical="center")

        ws1 = wb.active
        ws1.title = "Summary"
        ws1.sheet_view.showGridLines = False

        ws1.merge_cells("A1:C1")
        ws1["A1"] = f"Pair Analysis: {ticker_a} vs {ticker_b}"
        ws1["A1"].font = gold_font
        ws1["A1"].fill = dark_fill
        ws1["A1"].alignment = left
        ws1.row_dimensions[1].height = 32

        ws1.merge_cells("A2:C2")
        ws1["A2"] = f"Period: {period}   |   Rolling window: {window} days"
        ws1["A2"].font = muted_font
        ws1["A2"].fill = dark_fill
        ws1["A2"].alignment = left
        ws1.row_dimensions[2].height = 18

        ws1.merge_cells("A4:C4")
        ws1["A4"] = "PAIR STATISTICS"
        ws1["A4"].font = Font(color="C8A96E", bold=True, name="Courier New", size=10)
        ws1["A4"].fill = dark_fill
        ws1["A4"].alignment = left
        ws1.row_dimensions[4].height = 20

        pair_rows = [
            ("Metric", "Value"),
            ("Correlation", stats["correlation"]),
            ("R²", stats["r2"]),
            (f"Beta ({ticker_a}/{ticker_b})", stats["beta"]),
            ("Tracking Error (Ann.)", f"{stats['tracking_error']*100:.2f}%"),
        ]
        for i, (label, val) in enumerate(pair_rows, start=5):
            is_hdr = i == 5
            alt = i % 2 == 0
            for j, v in enumerate([label, val], start=1):
                cell = ws1.cell(row=i, column=j, value=v)
                cell.fill = header_fill if is_hdr else (alt_fill if alt else dark_fill)
                cell.font = dark_font if is_hdr else white_font
                cell.alignment = center
            ws1.row_dimensions[i].height = 20

        sr = 5 + len(pair_rows) + 2
        ws1.merge_cells(f"A{sr}:C{sr}")
        ws1[f"A{sr}"] = "INDIVIDUAL STOCK STATISTICS"
        ws1[f"A{sr}"].font = Font(color="C8A96E", bold=True, name="Courier New", size=10)
        ws1[f"A{sr}"].fill = dark_fill
        ws1[f"A{sr}"].alignment = left
        ws1.row_dimensions[sr].height = 20

        stock_rows = [
            ("Metric", ticker_a, ticker_b),
            ("Total Return", f"{stats['ticker_a']['total_return']*100:.2f}%", f"{stats['ticker_b']['total_return']*100:.2f}%"),
            ("Ann. Volatility", f"{stats['ticker_a']['ann_vol']*100:.2f}%", f"{stats['ticker_b']['ann_vol']*100:.2f}%"),
            ("Sharpe (rf=0)", f"{stats['ticker_a']['sharpe']:.2f}", f"{stats['ticker_b']['sharpe']:.2f}"),
            ("Max Drawdown", f"-{stats['ticker_a']['max_drawdown']*100:.2f}%", f"-{stats['ticker_b']['max_drawdown']*100:.2f}%"),
            ("Best Day", f"{stats['ticker_a']['best_day']*100:.2f}%", f"{stats['ticker_b']['best_day']*100:.2f}%"),
            ("Worst Day", f"{stats['ticker_a']['worst_day']*100:.2f}%", f"{stats['ticker_b']['worst_day']*100:.2f}%"),
        ]
        for i, row_data in enumerate(stock_rows, start=sr + 1):
            is_hdr = i == sr + 1
            alt = i % 2 == 0
            for j, v in enumerate(row_data, start=1):
                cell = ws1.cell(row=i, column=j, value=v)
                cell.fill = header_fill if is_hdr else (alt_fill if alt else dark_fill)
                cell.font = dark_font if is_hdr else white_font
                cell.alignment = center
            ws1.row_dimensions[i].height = 20

        for col, w in [(1, 28), (2, 16), (3, 16)]:
            ws1.column_dimensions[get_column_letter(col)].width = w

        ws2 = wb.create_sheet("Price Data")
        ws2.sheet_view.showGridLines = False

        idx_a   = [round(p / prices_a[0] * 100, 4) for p in prices_a]
        idx_b   = [round(p / prices_b[0] * 100, 4) for p in prices_b]
        idx_spy = [round(p / prices_spy[0] * 100, 4) for p in prices_spy]
        spread  = [round(a / b, 6) for a, b in zip(prices_a, prices_b)]

        headers = ["Date", ticker_a, ticker_b, "SPY",
                   f"{ticker_a} Idx", f"{ticker_b} Idx", "SPY Idx",
                   f"Spread ({ticker_a}/{ticker_b})", f"Rolling Corr ({window}d)"]
        for j, h in enumerate(headers, start=1):
            cell = ws2.cell(row=1, column=j, value=h)
            cell.fill = header_fill
            cell.font = dark_font
            cell.alignment = center
        ws2.row_dimensions[1].height = 22

        for i, date in enumerate(dates):
            rn = i + 2
            alt = i % 2 == 0
            row_vals = [
                date,
                round(prices_a[i], 4),
                round(prices_b[i], 4),
                round(prices_spy[i], 4),
                idx_a[i], idx_b[i], idx_spy[i],
                spread[i],
                rolling_corr[i] if i < len(rolling_corr) else None,
            ]
            for j, val in enumerate(row_vals, start=1):
                cell = ws2.cell(row=rn, column=j, value=val)
                cell.fill = alt_fill if alt else dark_fill
                cell.font = white_font
                cell.alignment = center
            ws2.row_dimensions[rn].height = 18

        for col, w in enumerate([14,12,12,12,12,12,12,22,20], start=1):
            ws2.column_dimensions[get_column_letter(col)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"pair_analysis_{ticker_a}_{ticker_b}_{period}.xlsx",
        )

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Export failed: {str(e)}"}), 500


@app.route("/")
def index():
    return send_from_directory(".", "index.html")


if __name__ == "__main__":
    print("\n✓ Pair Analysis server running at http://localhost:5000\n")
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
